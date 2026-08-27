"""公文/政策文档内容提取器 — PDF/DOCX/XLSX 解析 + LLM 降级"""

from __future__ import annotations

import logging
import re
import subprocess
import unicodedata
import zipfile
from pathlib import Path

logger = logging.getLogger(__name__)

MAX_PREVIEW_CHARS = 5000


def _strip_control_chars(text: str) -> str:
    """移除控制字符，保留换行和制表符"""
    return "".join(c for c in text if c == "\n" or c == "\t" or unicodedata.category(c)[0] != "C")


def _clean_title(filename: str) -> str:
    """从文件名清理出标题"""
    title = filename
    # 去除扩展名
    for ext in [".pdf", ".docx", ".doc", ".xlsx", ".md", ".txt"]:
        if title.lower().endswith(ext):
            title = title[: -len(ext)]
            break
    # 去除版本号 (v1, v2.1 等)
    title = re.sub(r"\s*v[\d.]+$", "", title)
    # 去除多余空白
    title = re.sub(r"\s+", " ", title).strip()
    return title or filename


_DOC_NUM_PATTERN = re.compile(r"[〔\[（][\w\d]+[号〕\]）]")


def _extract_doc_number(filename: str, content_hint: str = "") -> list[str]:
    """从文件名和内容中提取文号"""
    numbers = []
    # 从文件名提取
    for m in _DOC_NUM_PATTERN.finditer(filename):
        numbers.append(m.group())
    # 从内容中提取（可选）
    if content_hint:
        for m in _DOC_NUM_PATTERN.finditer(content_hint):
            numbers.append(m.group())
    seen: set[str] = set()
    result: list[str] = []
    for n in numbers:
        if n not in seen:
            seen.add(n)
            result.append(n)
    return result


def _try_pdf_extract(fp: Path) -> str | None:
    """尝试用 pdftotext 提取 PDF 内容"""
    try:
        result = subprocess.run(
            ["pdftotext", "-layout", "-nopgbrk", str(fp), "-"],
            capture_output=True,
            text=True,
            timeout=30,
        )
        if result.returncode == 0 and result.stdout.strip():
            return _strip_control_chars(result.stdout[:MAX_PREVIEW_CHARS])
        logger.warning("pdftotext returned empty for %s", fp.name)
    except FileNotFoundError:
        logger.info("pdftotext not installed, skipping PDF extraction")
    except subprocess.TimeoutExpired:
        logger.warning("pdftotext timeout for %s", fp.name)
    except Exception as e:
        logger.warning("pdftotext failed for %s: %s", fp.name, e)
    return None


def _try_docx_extract(fp: Path) -> str | None:
    """从 DOCX 中提取文本"""
    try:
        import docx  # type: ignore[import-not-found]

        doc = docx.Document(str(fp))
        texts = [p.text for p in doc.paragraphs if p.text.strip()]
        return _strip_control_chars(" | ".join(texts)[:MAX_PREVIEW_CHARS])
    except ImportError:
        pass
    # 降级：zipfile 读 document.xml
    try:
        import xml.etree.ElementTree as ET

        with zipfile.ZipFile(fp) as z:
            if "word/document.xml" in z.namelist():
                tree = ET.fromstring(z.read("word/document.xml"))
                texts = []
                for p in tree.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p"):
                    t = "".join(
                        ele.text or ""
                        for ele in p.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t")
                    )
                    if t.strip():
                        texts.append(t.strip())
                return _strip_control_chars(" | ".join(texts)[:MAX_PREVIEW_CHARS])
    except Exception:
        pass
    return None


def _try_xlsx_extract(fp: Path) -> str | None:
    """从 XLSX 中提取文本"""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        texts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(max_row=50, values_only=True):
                row_text = " | ".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    texts.append(row_text)
        return _strip_control_chars(" | ".join(texts)[:MAX_PREVIEW_CHARS])
    except ImportError:
        pass
    # 降级：zipfile 读 sharedStrings.xml
    try:
        with zipfile.ZipFile(fp) as z:
            sheets = [n for n in z.namelist() if n.startswith("xl/worksheets/sheet") and n.endswith(".xml")]
            if not sheets:
                return None
            # 读 shared strings
            # 安全解析：禁用外部实体防止 XXE
            try:
                from defusedxml.ElementTree import fromstring as _safe_fromstring

                sis_xml = z.read("xl/sharedStrings.xml") if "xl/sharedStrings.xml" in z.namelist() else b""
                sis = _safe_fromstring(sis_xml) if sis_xml else None
            except ImportError:
                import xml.etree.ElementTree as _et

                parser = _et.XMLParser()
                sis_xml = z.read("xl/sharedStrings.xml") if "xl/sharedStrings.xml" in z.namelist() else b""
                sis = _et.fromstring(sis_xml, parser=parser) if sis_xml else None
            ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            texts = []
            if sis is not None:
                for si in sis.findall(".//s:si", ns):
                    t = "".join(ele.text or "" for ele in si.iter())
                    if t.strip():
                        texts.append(t.strip())
            return _strip_control_chars(" | ".join(texts)[:MAX_PREVIEW_CHARS])
    except Exception:
        pass
    return None


def _try_llm_fallback(fp: Path, ext: str) -> str | None:
    """用本地 LLM 解析非标准格式（可选依赖，降级返回 None）"""
    # kronos 是独立项目，不直接依赖
    # 如需 LLM 解析：pip install -e /path/to/kronos
    return None


def extract_file_content(fp: Path) -> str | None:
    """提取文件内容（自动识别格式：PDF/DOCX/XLSX/其他）"""
    ext = fp.suffix.lower()

    # PDF
    if ext == ".pdf":
        result = _try_pdf_extract(fp)
        if result:
            return result

    # DOCX
    elif ext == ".docx":
        result = _try_docx_extract(fp)
        if result:
            return result

    # XLSX
    elif ext == ".xlsx":
        result = _try_xlsx_extract(fp)
        if result:
            return result

    # 其他文件（md/txt 等）
    try:
        text = fp.read_text(encoding="utf-8", errors="ignore")[:MAX_PREVIEW_CHARS]
        if text.strip():
            return _strip_control_chars(text)
    except Exception:
        pass

    # 最后用 LLM 降级
    return _try_llm_fallback(fp, ext)
